from openpyxl import Workbook
from django.http import HttpResponse
from menu.models import MenuItem
from order.models import Order, OrderItem
from customer.models import Customer
from .models import Staff


def create_orders_sheet():
    workbook = Workbook()
    orders_sheet = workbook.active
    orders_sheet.title = "Order items"
    orders_sheet.append(
        [
            "Id",
            "Quantity",
            "Subtotal",
            "Customer Phone Number",
            "Order Date",
            "Table Number",
        ]
    )

    order_items = OrderItem.objects.all()
    for item in order_items:
        order_date_naive = (
            item.order.order_date.replace(tzinfo=None)
            if item.order.order_date
            else None
        )
        phone_number = item.order.customer.phone_number if item.order.customer else ""
        orders_sheet.append(
            [
                item.id,
                item.quantity,
                item.subtotal,
                phone_number,
                order_date_naive,
                item.order.table_number,
            ]
        )
    return workbook


def create_customers_sheet():
    workbook = Workbook()
    customers_sheet = workbook.active
    customers_sheet.title = "Customers"
    customers_sheet.append(
        [
            "Customer ID",
            "First Name",
            "Last Name",
            "Phone Number",
            "Number of Orders",
            "Total Amount Paid",
            "Points",
            "Last Order Date",
        ]
    )

    customers = Customer.objects.all()
    for customer in customers:
        orders_by_customer = Order.objects.filter(customer=customer)
        number_of_orders = orders_by_customer.count()
        total_amount_paid = sum(order.total_price for order in orders_by_customer)
        last_order_date = (
            orders_by_customer.order_by("-order_date")
            .first()
            .order_date.replace(tzinfo=None)
            if orders_by_customer.exists()
            else None
        )
        customers_sheet.append(
            [
                customer.id,
                customer.first_name if hasattr(customer, "first_name") else "",
                customer.last_name if hasattr(customer, "last_name") else "",
                customer.phone_number,
                number_of_orders,
                total_amount_paid,
                customer.points,
                last_order_date,
            ]
        )
    return workbook


def create_staff_sheet():
    workbook = Workbook()
    staff_sheet = workbook.active
    staff_sheet.title = "Staff"
    staff_sheet.append(
        [
            "Staff ID",
            "First Name",
            "Last Name",
            "Phone Number",
            "Number of Orders",
        ]
    )

    staff_members = Staff.objects.all()
    for staff in staff_members:
        orders_by_staff = Order.objects.filter(staff=staff)
        number_of_orders = orders_by_staff.count()
        staff_sheet.append(
            [
                staff.pk,
                staff.first_name if hasattr(staff, "first_name") else "",
                staff.last_name if hasattr(staff, "last_name") else "",
                staff.phone_number,
                number_of_orders,
            ]
        )
    return workbook


def create_menu_items_sheet():
    workbook = Workbook()
    menu_items_sheet = workbook.active
    menu_items_sheet.title = "Menu Items"
    menu_items_sheet.append(
        [
            "Menu Item ID",
            "Name",
            "Price",
            "Points",
            "Category",
        ]
    )

    menu_items = MenuItem.objects.all()
    for item in menu_items:
        category_name = item.category.name if item.category else ""
        menu_items_sheet.append(
            [
                item.id,
                item.name,
                item.price,
                item.points,
                category_name,
            ]
        )
    return workbook


def generate_excel_response(workbook, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
